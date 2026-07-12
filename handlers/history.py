import os
import logging
from datetime import datetime
from aiogram import Router, F, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile

from utils.keyboards import get_main_keyboard, get_github_profile_actions
from utils.excel_generator import create_excel_file_fns, create_excel_file_phone
from utils.formatters import format_github_profile, format_github_compare
from utils.database import get_user_history, get_request_results
from core import bot

logger = logging.getLogger(__name__)
router = Router()


@router.message(Command("history"))
async def cmd_history(message: types.Message, state: FSMContext):
    """Обработчик команды /history - показывает историю запросов пользователя."""
    await state.clear()
    await show_history(message.from_user.id, message)


@router.callback_query(F.data == "open_history")
async def handle_open_history_button(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик кнопки 'История запросов' из главного меню."""
    await callback.answer()
    await state.clear()
    try:
        await callback.message.delete()
    except:
        pass
    await show_history(callback.from_user.id, callback.message)


async def show_history(user_id: int, message: types.Message, offset: int = 0):
    """
    Показывает историю запросов пользователя с пагинацией.
    """
    limit = 5
    history = get_user_history(user_id, limit=limit + 1, offset=offset)

    if not history:
        await message.answer(
            "📜 *История запросов пуста*\n\n"
            "Вы ещё не выполняли поисков. Начните с главного меню!",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )
        return

    has_more = len(history) > limit
    if has_more:
        history = history[:limit]

    text_lines = [
        "📜 *Ваши последние запросы:*\n\n"
        "📖 *Легенда:*",
        "✅ — есть результат, можно посмотреть",
        "📭 — ничего не нашли (чисто, не ошибка)",
        "⛔ — отменен",
        "❌ — что-то пошло не так\n",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    ]

    keyboard_buttons = []

    for req in history:
        request_id = req['request_id']
        search_type = req['search_type']
        search_value = req['search_value']
        results_count = req['results_count']
        status = req['status']
        created_at_raw = req['created_at']

        # Парсим дату и время
        try:
            created_dt = datetime.fromisoformat(created_at_raw)
            created_formatted = created_dt.strftime("%Y-%m-%d | %H.%M.%S")
        except:
            created_formatted = created_at_raw[:10]

        # Логика выбора эмодзи по статусу
        if status == 'cancelled':
            status_icon = "⛔"  # Отменено пользователем
        elif status == 'completed' and results_count > 0:
            status_icon = "✅"  # Успешно найдено
        elif status == 'not_found' or (status == 'completed' and results_count == 0):
            status_icon = "📭"  # Пусто (ничего не найдено)
        else:
            status_icon = "❌"  # Настоящая ошибка

        # Иконка типа поиска
        if search_type == 'fio':
            icon = "👤"
            type_name = "ФИО"
        elif search_type == 'phone':
            icon = "📱"
            type_name = "Телефон"
        elif search_type == 'github':
            icon = "🐙"
            type_name = "GitHub"
        elif search_type == 'github_compare':
            icon = "⚔️"
            type_name = "GitHub сравнение"
        else:
            icon = "🔍"
            type_name = search_type

        # Формируем текст кнопки
        display_value = search_value if len(search_value) <= 25 else search_value[:22] + "..."
        button_text = f"{icon} {display_value} ({created_formatted})"

        # Добавляем в текст сообщения
        text_lines.append(f"• {icon} *{type_name}:* `{search_value}` | 📅 {created_formatted} | {status_icon}")

        keyboard_buttons.append([
            InlineKeyboardButton(
                text=button_text,
                callback_data=f"history_download_{request_id}"
            )
        ])

    keyboard_buttons.append([
        InlineKeyboardButton(
            text="📥 Скачать всю историю в Excel",
            callback_data="history_export_all"
        )
    ])

    # Кнопки пагинации
    pagination_buttons = []
    if offset > 0:
        pagination_buttons.append(
            InlineKeyboardButton(text="⬅️ Назад", callback_data=f"history_page_{offset - limit}")
        )
    if has_more:
        pagination_buttons.append(
            InlineKeyboardButton(text="Вперёд ➡️", callback_data=f"history_page_{offset + limit}")
        )

    if pagination_buttons:
        keyboard_buttons.append(pagination_buttons)

    # Кнопка возврата в меню
    keyboard_buttons.append([
        InlineKeyboardButton(text="🏠 В главное меню", callback_data="back_to_menu")
    ])

    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)

    await message.answer(
        "\n".join(text_lines),
        reply_markup=keyboard,
        parse_mode="Markdown"
    )


@router.callback_query(F.data.startswith("history_page_"))
async def handle_history_pagination(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик пагинации истории запросов."""
    await callback.answer()
    offset = int(callback.data.replace("history_page_", ""))

    try:
        await callback.message.delete()
    except:
        pass

    await show_history(callback.from_user.id, callback.message, offset=offset)


@router.callback_query(F.data.startswith("history_download_"))
async def handle_history_download(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик скачивания результата из истории запросов."""
    await callback.answer("⏳ Загружаю данные...")

    request_id = int(callback.data.replace("history_download_", ""))
    request_data = get_request_results(request_id)

    if not request_data:
        await callback.message.answer("❌ Данные запроса не найдены в базе")
        return

    search_type = request_data['search_type']
    search_value = request_data['search_value']
    results = request_data['results']
    status = request_data['status']  # Получаем статус

    # Парсим дату исходного запроса
    try:
        created_dt = datetime.fromisoformat(request_data['created_at'])
        created_formatted = created_dt.strftime("%Y-%m-%d | %H.%M.%S")
    except:
        created_formatted = request_data['created_at'][:10]

    # Обработка отменённых запросов
    if status == 'cancelled':
        await callback.message.answer(
            f"⛔ *Запрос был отменён*\n\n"
            f"🔹 Тип поиска: `{search_type}`\n"
            f"🔹 Значение: `{search_value}`\n"
            f"📅 _Дата запроса: {created_formatted}_\n\n",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )
        return

    # Обработка пустых результатов и ошибок
    if not results or len(results) == 0:
        # Определяем эмодзи и текст в зависимости от статуса
        if status in ('not_found', 'completed'):
            # Поиск прошёл успешно, но ничего не нашли
            status_icon = "📭"
            title = "Нет данных по запросу"
        elif status in ('error', 'parser_error'):
            # Настоящая ошибка (сеть, API, сбой)
            status_icon = "❌"
            title = "Ошибка при запросе"
        else:
            status_icon = "❌"
            title = "Нет данных"

        await callback.message.answer(
            f"{status_icon} *{title}*\n\n"
            f"🔹 Тип: `{search_type}`\n"
            f"🔹 Значение: `{search_value}`\n"
            f"📅 _Дата запроса: {created_formatted}_\n\n",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )
        return

    # ==========================================
    # 🐙 GITHUB: выводим профиль с inline-кнопками
    # ==========================================
    if search_type == 'github':
        data = results[0] if isinstance(results, list) else results
        user = data.get("user")
        repos_stats = data.get("repos_stats")

        if not user:
            await callback.message.answer("❌ Данные профиля повреждены в базе")
            return

        profile_text = format_github_profile(user, repos_stats)
        full_text = f"{profile_text}\n\n📅 _Запрос из истории: {created_formatted}_"
        username = user.get("login")

        await callback.message.answer(
            full_text,
            parse_mode="Markdown",
            reply_markup=get_github_profile_actions(username),
            disable_web_page_preview=True
        )

    # ==========================================
    # ⚔️ GITHUB COMPARE: выводим сравнение
    # ==========================================
    elif search_type == 'github_compare':
        data = results[0] if isinstance(results, list) else results
        user1 = data.get("user1")
        stats1 = data.get("stats1")
        user2 = data.get("user2")
        stats2 = data.get("stats2")

        if not all([user1, user2]):
            await callback.message.answer("❌ Данные сравнения повреждены в базе")
            return

        compare_text = format_github_compare(user1, stats1, user2, stats2)
        full_text = f"{compare_text}\n\n📅 _Запрос из истории: {created_formatted}_"

        await callback.message.answer(
            full_text,
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )

    # ==========================================
    # 👤 ФИО: генерируем Excel-файл
    # ==========================================
    elif search_type == 'fio':
        filepath = create_excel_file_fns(search_value, results)
        caption = (
            f"📁 *Результаты из истории: ФИО {search_value}*\n"
            f"🔹 Найдено записей: *{len(results)}*\n"
            f"🔹 Источник: ФНС России (list-org.com)\n"
            f"🔹 📅 Дата исходного запроса: {created_formatted}"
        )

        await bot.send_chat_action(chat_id=callback.message.chat.id, action="upload_document")
        await callback.message.answer_document(
            document=FSInputFile(filepath),
            caption=caption,
            parse_mode="Markdown"
        )

        try:
            os.remove(filepath)
        except Exception as e:
            logger.error(f"Не удалось удалить файл: {e}")

    # ==========================================
    # 📱 ТЕЛЕФОН: генерируем Excel-файл
    # ==========================================
    elif search_type == 'phone':
        filepath = create_excel_file_phone(search_value, results)
        caption = (
            f"📁 *Результаты из истории: Телефон {search_value}*\n"
            f"🔹 Найдено организаций: *{len(results)}*\n"
            f"🔹 Источник: list-org.com\n"
            f"🔹 📅 Дата исходного запроса: {created_formatted}"
        )

        await bot.send_chat_action(chat_id=callback.message.chat.id, action="upload_document")
        await callback.message.answer_document(
            document=FSInputFile(filepath),
            caption=caption,
            parse_mode="Markdown"
        )

        try:
            os.remove(filepath)
        except Exception as e:
            logger.error(f"Не удалось удалить файл: {e}")

    else:
        await callback.message.answer(f"❌ Неизвестный тип запроса: {search_type}")


@router.callback_query(F.data == "history_export_all")
async def handle_history_export_all(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик экспорта всей истории в один Excel-файл."""
    await callback.answer("⏳ Формирую полную историю...")

    history = get_user_history(callback.from_user.id, limit=1000)

    if not history:
        await callback.message.answer("📜 История пуста")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "История запросов"

    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    ws['A1'].value = "История поисковых запросов"
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['№', 'Тип поиска', 'Запрос', 'Результатов', 'Дата и время', 'Статус']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Маппинг статусов
    status_map = {
        'completed': '✅ Выполнен',
        'error': '❌ Ошибка',
        'parser_error': '❌ Ошибка',
        'not_found': '📭 Ничего не найдено',
        'cancelled': '⛔ Отменён'
    }

    column_widths = [5, 20, 40, 12, 25, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_num, req in enumerate(history, 4):
        search_type_map = {
            'fio': '👤 ФИО',
            'phone': '📱 Телефон',
            'github': '🐙 GitHub',
            'github_compare': '⚔️ GitHub сравнение',
            'plate': '🚗 Госномер'
        }
        search_type_text = search_type_map.get(req['search_type'], req['search_type'])

        # Логика выбора статуса
        # СНАЧАЛА проверяем отмены
        if req['status'] == 'cancelled':
            status_text = '⛔ Отменён'
        # ПОТОМ проверяем пустые результаты (completed с 0 результатов)
        elif req['status'] == 'completed' and req['results_count'] == 0:
            status_text = '📭 Ничего не найдено'
        # В остальных случаях используем маппинг
        else:
            status_text = status_map.get(req['status'], req['status'])

        # Парсим дату и время
        try:
            created_dt = datetime.fromisoformat(req['created_at'])
            created_formatted = created_dt.strftime("%Y-%m-%d | %H.%M.%S")
        except:
            created_formatted = req['created_at'][:10]

        ws.cell(row=row_num, column=1, value=row_num - 3).border = thin_border
        ws.cell(row=row_num, column=2, value=search_type_text).border = thin_border
        ws.cell(row=row_num, column=3, value=req['search_value']).border = thin_border
        ws.cell(row=row_num, column=4, value=req['results_count']).border = thin_border
        ws.cell(row=row_num, column=5, value=created_formatted).border = thin_border
        ws.cell(row=row_num, column=6, value=status_text).border = thin_border

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"История_запросов_{timestamp}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)

    caption = (
        f"📜 *Полная история запросов*\n"
        f"🔹 Всего запросов: *{len(history)}*\n"
        f"🔹 Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )

    await bot.send_chat_action(chat_id=callback.message.chat.id, action="upload_document")
    await callback.message.answer_document(
        document=FSInputFile(filepath),
        caption=caption,
        parse_mode="Markdown"
    )

    try:
        os.remove(filepath)
    except Exception as e:
        logger.error(f"Не удалось удалить файл: {e}")