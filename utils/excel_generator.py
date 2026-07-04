import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_file_fns(fio: str, results: list) -> str:
    """Создает Excel-файл с результатами поиска в ЕГРЮЛ/ЕГРИП."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты поиска ФНС"

    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    ws['A1'].value = f"Результаты поиска в ЕГРЮЛ/ЕГРИП по ФИО: {fio}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:I2')
    ws['A2'].value = f"Дата поиска: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws['A2'].font = Font(name='Arial', italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:I3')
    ws['A3'].value = "Источник: Федеральная налоговая служба (ФНС России) - list-org.com"
    ws['A3'].font = Font(name='Arial', italic=True, size=10, color='555555')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    headers = ['№', 'Название', 'Тип', 'Статус', 'ИНН', 'КПП', 'Юридический адрес', 'Роль',
               'Для просмотра доп. информации']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    column_widths = [5, 40, 10, 15, 18, 15, 50, 30, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_num, result in enumerate(results, 6):
        ws.cell(row=row_num, column=1, value=row_num - 5).border = thin_border
        ws.cell(row=row_num, column=2, value=result['name']).border = thin_border
        ws.cell(row=row_num, column=3, value=result['type']).border = thin_border
        ws.cell(row=row_num, column=4, value=result['status']).border = thin_border
        ws.cell(row=row_num, column=5, value=result['inn']).border = thin_border
        ws.cell(row=row_num, column=6, value=result['kpp']).border = thin_border
        ws.cell(row=row_num, column=7, value=result['address']).border = thin_border
        ws.cell(row=row_num, column=8, value=result['role']).border = thin_border
        link_cell = ws.cell(row=row_num, column=9, value=result['link'])
        link_cell.hyperlink = result['link']
        link_cell.font = Font(color='0565C2', underline='single')
        link_cell.border = thin_border

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"ФНС_{fio.replace(' ', '_')}_{timestamp}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    logging.info(f"Excel-файл сохранен: {filepath}")
    return filepath

def create_excel_file_phone(phone: str, results: list) -> str:
    """Создает Excel-файл с результатами поиска организаций по телефону."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты поиска по телефону"

    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:I1')
    ws['A1'].value = f"Результаты поиска организаций по телефону: {phone}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:I2')
    ws['A2'].value = f"Дата поиска: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws['A2'].font = Font(name='Arial', italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:I3')
    ws['A3'].value = "Источник: list-org.com"
    ws['A3'].font = Font(name='Arial', italic=True, size=10, color='555555')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    headers = ['№', 'Название', 'Тип', 'Статус', 'ИНН', 'КПП', 'Юридический адрес', 'Роль',
               'Для просмотра доп. информации']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    column_widths = [5, 40, 10, 15, 18, 15, 50, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_num, result in enumerate(results, 6):
        ws.cell(row=row_num, column=1, value=row_num - 5).border = thin_border
        ws.cell(row=row_num, column=2, value=result['name']).border = thin_border
        ws.cell(row=row_num, column=3, value=result['type']).border = thin_border
        ws.cell(row=row_num, column=4, value=result['status']).border = thin_border
        ws.cell(row=row_num, column=5, value=result['inn']).border = thin_border
        ws.cell(row=row_num, column=6, value=result['kpp']).border = thin_border
        ws.cell(row=row_num, column=7, value=result['address']).border = thin_border
        ws.cell(row=row_num, column=8, value=result['role']).border = thin_border
        link_cell = ws.cell(row=row_num, column=9, value=result['link'])
        link_cell.hyperlink = result['link']
        link_cell.font = Font(color='0565C2', underline='single')
        link_cell.border = thin_border

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    filename = f"Телефон_{phone.replace('+', '').replace(' ', '_')}_{timestamp}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    logging.info(f"Excel-файл сохранен: {filepath}")
    return filepath